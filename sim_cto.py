# sim_cto.py
import io
from pathlib import Path
from contextlib import redirect_stderr, redirect_stdout

import pandas as pd
import streamlit as st
import pulp as pl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from style import aplicar_estilo_personalizado

# Módulo do modelo importado como módulo (para reatribuir PROJECT_DIR sem editar o arquivo).
import dados_ferias_cto
from modelo_ferias_cto import resolver_modelo_tatico
from resultados_tatico import extrair_resultados_tatico

# Parâmetros fixos do solver (sem controles na tela).
TIME_LIMIT = 300   # segundos
GAP = 0.05         # 5%
NOME_BASELINE_LOCAL = "Baseline_Resultados_Tatico.xlsx"


def usando_google():
    """True quando o deploy possui credenciais e IDs configurados."""
    try:
        tem_credencial = (
            "google_credentials_json" in st.secrets
            or "gcp_service_account" in st.secrets
        )
        return tem_credencial and "planilhas" in st.secrets
    except Exception:
        return False


def limpar_cache():
    if usando_google():
        from conexao_google import limpar_cache as _limpar
        _limpar()


# =========================================================================
#                 CARGA DE DADOS DO MODELO (CACHEADA)
# =========================================================================
@st.cache_data(show_spinner=False)
def carregar_solicitacoes_pendentes_base():
    """Carrega as opcoes exibidas no seletor de solicitacoes pendentes."""
    if usando_google():
        from conexao_google import baixar_planilhas
        dados_ferias_cto.PROJECT_DIR = Path(baixar_planilhas())
    return dados_ferias_cto.carregar_solicitacoes_ferias_pendentes()


@st.cache_data(show_spinner=False)
def carregar_dados_base(solicitacoes_aprovadas_teste=()):
    """Carrega os dados do modelo.

    Nuvem: exporta as planilhas para uma pasta temporaria do servidor Streamlit.
    Local: usa o PROJECT_DIR que o próprio modelo já define (a pasta dos .py).
    """
    if usando_google():
        from conexao_google import baixar_planilhas
        dados_ferias_cto.PROJECT_DIR = Path(baixar_planilhas())
    solicitacoes = [
        {"matricula": matricula, "inicio": inicio, "fim": fim}
        for matricula, inicio, fim in solicitacoes_aprovadas_teste
    ]
    log = io.StringIO()
    with redirect_stdout(log):
        dados = dados_ferias_cto.carregar_dados(
            solicitacoes_aprovadas_teste=solicitacoes
        )
    return dados, log.getvalue()


def extrair_aviso_distancia(log):
    """Extrai do log o bloco de aviso sobre pares de cidades ausentes na aba Distancias."""
    linhas = log.splitlines()
    bloco, capturando = [], False
    for ln in linhas:
        if ln.startswith("ATENCAO:") and "aba" in ln:
            capturando = True
        if capturando:
            if ln.strip() == "" and bloco:
                break
            bloco.append(ln)
    return "\n".join(bloco).strip()


# =========================================================================
#                 EXECUÇÃO DO MODELO E RESULTADOS
# =========================================================================
def rodar_modelo(dados, time_limit, gap, log_inicial=""):
    log = io.StringIO()
    if log_inicial:
        log.write(log_inicial.rstrip() + "\n\n")
    with redirect_stdout(log), redirect_stderr(log):
        print("Iniciando resolucao do modelo tatico...")
        model, status, variaveis = resolver_modelo_tatico(
            dados, time_limit=time_limit, gap=gap, config={}
        )
        print("Extraindo tabelas de resultado...")
        resultados = extrair_resultados_tatico(dados, model, status, variaveis)
        print("Resultados extraidos.")
    return resultados, pl.LpStatus[status], log.getvalue()


def _reais(v):
    try:
        s = f"{abs(float(v)):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return str(v)
    return f"-R$ {s}" if float(v) < 0 else f"R$ {s}"


# Cabeçalhos intuitivos por tabela (aplicados na exibição e no Excel).
RENOMES = {
    "resumo": {"indicador": "Indicador", "valor": "Valor"},
    "demanda": {
        "cargo": "Cargo", "demanda": "Demanda (postos-dia)", "coberta": "Coberta",
        "nao_atendida": "Não atendida", "receita_coberta": "Receita coberta",
        "salarios_IA_IE": "Salários fixos (IA/IE)", "contratacao_IS": "Contratação de suplentes",
        "adicional_noturno": "Adicional noturno", "transporte_mudanca": "Transporte por mudança",
    },
    "contratacoes": {
        "suplente": "Matrícula", "cargo": "Cargo", "cidade": "Cidade", "turno_original": "Turno",
        "flex": "Disponibilidade", "quantidade_contratada": "Suplentes contratados",
        "custo_total": "Custo total",
    },
    "faltas": {
        "colaborador": "Matrícula em férias", "projeto": "Projeto", "cargo": "Cargo",
        "data_inicio": "Data inicial", "data_fim": "Data final", "dias": "Quantidade de dias",
        "turno": "Turno",
        "falta": "Postos-dia não atendidos", "receita_nao_gerada": "Receita não gerada",
    },
    "alocacoes": {
        "tipo": "Tipo", "suplente": "Matrícula", "nome_suplente": "Nome",
        "rota_id": "Rota", "colaborador_substituido": "Matrícula substituída",
        "projeto": "Projeto", "cargo_demanda": "Cargo coberto",
        "data_inicio": "Data inicial", "data_fim": "Data final",
        "turno": "Turno", "quantidade_dias": "Quantidade de dias",
    },
    "transporte": {
        "tipo": "Tipo", "suplente": "Matrícula", "rota_id": "Rota",
        "origem": "Origem", "sequencia_projetos": "Sequência de projetos",
        "periodo_rota": "Período da rota", "qtd_tarefas": "Quantidade de tarefas",
        "custo_transporte": "Custo de transporte",
        "custo_mobilizacao": "Custo de mobilização",
        "custo_noturno": "Adicional noturno", "custo_total_rota": "Custo total da rota",
        "projetos_mobilizados_cobrados": "Projetos mobilizados cobrados",
        "pernas_transporte": "Detalhamento do transporte",
    },
}


INDICADORES_MONETARIOS_RESUMO = {
    "Custo decisório", "Receita potencial total", "Receita não gerada", "Receita dos postos cobertos",
    "Contratação IS", "Adicional noturno", "Transporte por rota", "Mobilização",
    "Receita total gerada",
}

COLUNAS_MONETARIAS = {
    "demanda": {
        "receita_coberta", "salarios_IA_IE", "contratacao_IS",
        "adicional_noturno", "transporte_mudanca",
    },
    "contratacoes": {"custo_total"},
    "faltas": {"receita_nao_gerada"},
    "transporte": {
        "custo_transporte", "custo_mobilizacao", "custo_noturno", "custo_total_rota",
    },
}

FORMATO_MOEDA_EXCEL = '"R$" #,##0;-"R$" #,##0'


def _mostrar(df, tabela):
    """Remove colunas internas e aplica cabeçalhos intuitivos para exibição/Excel."""
    df2 = df.drop(columns=[c for c in ("t",) if c in df.columns], errors="ignore")
    df2 = df2.rename(columns=RENOMES.get(tabela, {})).copy()
    def texto_arrow(v):
        if v is None:
            return ""
        if isinstance(v, (list, tuple, dict, set)):
            return str(v)
        return "" if pd.isna(v) else str(v)
    for col in df2.select_dtypes(include=["object"]).columns:
        df2[col] = df2[col].map(texto_arrow)
    return df2


def _mostrar_streamlit(df, tabela):
    """Versao apenas visual, com todos os campos monetarios exibidos em reais."""
    df2 = _mostrar(df, tabela)
    if tabela == "resumo" and not df2.empty:
        df2 = df2.copy()
        valores_exibicao = []
        for indicador, valor in zip(df2["Indicador"], df2["Valor"]):
            if indicador in INDICADORES_MONETARIOS_RESUMO:
                valores_exibicao.append(_reais(valor))
                continue
            if valor is None or pd.isna(valor):
                valores_exibicao.append("")
                continue
            try:
                valores_exibicao.append(f"{float(valor):,.0f}".replace(",", "."))
            except (TypeError, ValueError):
                valores_exibicao.append(str(valor))
        # Toda a coluna visual passa a ter um unico tipo textual. Isso evita
        # inserir strings como "R$ ..." dentro da coluna numerica original.
        df2["Valor"] = pd.Series(valores_exibicao, index=df2.index, dtype="string")
        return df2

    for coluna_interna in COLUNAS_MONETARIAS.get(tabela, set()):
        coluna_exibida = RENOMES.get(tabela, {}).get(coluna_interna, coluna_interna)
        if coluna_exibida in df2.columns:
            df2[coluna_exibida] = df2[coluna_exibida].map(_reais)
    return df2


def _formatar_aba_excel(ws, tabela):
    """Aplica formatos numericos reais e uma largura legivel na aba exportada."""
    cabecalhos = {cell.value: cell.column for cell in ws[1]}

    if tabela == "resumo":
        for linha in range(2, ws.max_row + 1):
            indicador = ws.cell(linha, 1).value
            valor = ws.cell(linha, 2)
            if indicador in INDICADORES_MONETARIOS_RESUMO:
                valor.number_format = FORMATO_MOEDA_EXCEL
            elif indicador == "Mobilizações cobradas":
                valor.number_format = "#,##0"
    else:
        for coluna_interna in COLUNAS_MONETARIAS.get(tabela, set()):
            coluna_exibida = RENOMES.get(tabela, {}).get(coluna_interna, coluna_interna)
            indice = cabecalhos.get(coluna_exibida)
            if indice is None:
                continue
            for linha in range(2, ws.max_row + 1):
                ws.cell(linha, indice).number_format = FORMATO_MOEDA_EXCEL

    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 0:
        ws.auto_filter.ref = ws.dimensions
    for indice in range(1, ws.max_column + 1):
        valores = [ws.cell(linha, indice).value for linha in range(1, min(ws.max_row, 100) + 1)]
        largura = max((len(str(valor)) for valor in valores if valor is not None), default=10) + 2
        ws.column_dimensions[get_column_letter(indice)].width = min(max(largura, 12), 45)

    if tabela == "ferias":
        indice_cobertura = cabecalhos.get("Quem cobre e períodos")
        if indice_cobertura is not None:
            ws.column_dimensions[get_column_letter(indice_cobertura)].width = 70
            for linha in range(2, ws.max_row + 1):
                celula = ws.cell(linha, indice_cobertura)
                celula.alignment = Alignment(wrap_text=True, vertical="top")
                qtd_linhas = max(1, str(celula.value or "").count("\n") + 1)
                ws.row_dimensions[linha].height = max(18, 15 * qtd_linhas)


def metricas(resultados, status_str):
    contr = resultados["contratacoes"]
    faltas = resultados["faltas"]
    resumo = resultados.get("resumo", pd.DataFrame())
    mapa_resumo = (
        dict(zip(resumo["indicador"], resumo["valor"]))
        if {"indicador", "valor"}.issubset(resumo.columns)
        else {}
    )

    def valor_resumo(indicador, absoluto=False):
        valor = float(mapa_resumo.get(indicador, 0.0) or 0.0)
        return abs(valor) if absoluto else valor

    return {
        "status": status_str,
        "fo": float(resultados.get("custo_total_valor", float("nan"))),
        "lucro": float(resultados.get("lucro_calculado", float("nan"))),
        "receita_nao_gerada": valor_resumo("Receita não gerada", absoluto=True),
        "custo_contratacao": valor_resumo("Contratação IS", absoluto=True),
        "custo_transporte": valor_resumo("Transporte por rota", absoluto=True),
        "custo_mobilizacao": valor_resumo("Mobilização", absoluto=True),
        "mobilizacoes": valor_resumo("Mobilizações cobradas"),
        "custo_noturno": valor_resumo("Adicional noturno", absoluto=True),
        "contratacoes": float(contr["quantidade_contratada"].sum()) if not contr.empty else 0.0,
        "faltas": float(faltas["falta"].sum()) if not faltas.empty else 0.0,
    }


def montar_excel_resultados(resultados):
    buffer = io.BytesIO()
    abas = [
        ("Resumo", "resumo"), ("Demanda", "demanda"), ("Contratacoes", "contratacoes"),
        ("Ferias", "ferias"), ("Faltas", "faltas"), ("Alocacoes", "alocacoes"),
        ("Transporte", "transporte"),
    ]
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nome_aba, chave in abas:
            if chave in resultados:
                _mostrar(resultados[chave], chave).to_excel(writer, sheet_name=nome_aba, index=False)
                _formatar_aba_excel(writer.book[nome_aba], chave)
    buffer.seek(0)
    return buffer.getvalue()


def montar_abas_google(resultados, modo="baseline", comparacao=None):
    """Monta as tabelas que serao gravadas na planilha Google Resultados."""
    definicoes = [
        ("Resumo", "resumo"), ("Demanda", "demanda"),
        ("Contratacoes", "contratacoes"), ("Ferias", "ferias"),
        ("Faltas", "faltas"), ("Alocacoes", "alocacoes"),
        ("Transporte", "transporte"),
    ]
    prefixo = "" if modo == "baseline" else "Simulacao_"
    abas = {
        prefixo + nome_aba: _mostrar(resultados[chave], chave)
        for nome_aba, chave in definicoes
        if chave in resultados
    }
    if modo == "simulacao" and comparacao is not None:
        abas["Simulacao_Comparacao"] = comparacao.copy()
    return abas


def caminho_baseline_local():
    """Ponto unico para trocar futuramente o armazenamento local pelo Google Drive."""
    return Path(dados_ferias_cto.PROJECT_DIR) / NOME_BASELINE_LOCAL


def salvar_baseline_local(conteudo_xlsx):
    caminho = caminho_baseline_local()
    caminho.write_bytes(conteudo_xlsx)
    return caminho


def salvar_resultados_persistentes(resultados, modo, comparacao=None, conteudo_xlsx=None):
    """Salva na planilha Google; mantem fallback local apenas para desenvolvimento."""
    if usando_google():
        from conexao_google import salvar_abas_resultados

        abas = montar_abas_google(resultados, modo=modo, comparacao=comparacao)
        salvar_abas_resultados(abas)
        return "Planilha Google Resultados"
    if modo == "baseline" and conteudo_xlsx is not None:
        return str(salvar_baseline_local(conteudo_xlsx))
    return None


def carregar_metricas_baseline_local():
    if usando_google():
        from conexao_google import ler_aba_resultados

        resumo = ler_aba_resultados("Resumo")
        if resumo is None or resumo.empty:
            return None
        contratacoes = ler_aba_resultados("Contratacoes")
        faltas = ler_aba_resultados("Faltas")
        contratacoes = contratacoes if contratacoes is not None else pd.DataFrame()
        faltas = faltas if faltas is not None else pd.DataFrame()
    else:
        caminho = caminho_baseline_local()
        if not caminho.exists():
            return None
        # A comparacao usa as abas de negocio do proprio baseline; nao e necessario
        # manter uma aba tecnica exclusiva para metricas do cenario.
        resumo = pd.read_excel(caminho, sheet_name="Resumo")
        contratacoes = pd.read_excel(caminho, sheet_name="Contratacoes")
        faltas = pd.read_excel(caminho, sheet_name="Faltas")

    def numero(valor):
        if valor is None or pd.isna(valor):
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)
        texto = str(valor).strip().replace("R$", "").replace(" ", "")
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif texto.count(".") > 1 or ("." in texto and len(texto.rsplit(".", 1)[1]) == 3):
            texto = texto.replace(".", "")
        try:
            return float(texto)
        except ValueError:
            return 0.0

    coluna_indicador = next((c for c in ("Indicador", "indicador") if c in resumo.columns), None)
    coluna_valor = next((c for c in ("Valor", "valor") if c in resumo.columns), None)
    mapa_resumo = {}
    if coluna_indicador and coluna_valor:
        mapa_resumo = {
            str(indicador): numero(valor)
            for indicador, valor in zip(resumo[coluna_indicador], resumo[coluna_valor])
        }

    def valor_resumo(indicador, absoluto=False):
        valor = float(mapa_resumo.get(indicador, 0.0))
        return abs(valor) if absoluto else valor

    coluna_contratacoes = next(
        (c for c in ("Suplentes contratados", "quantidade_contratada") if c in contratacoes.columns),
        None,
    )
    coluna_faltas = next(
        (c for c in ("Postos-dia não atendidos", "falta") if c in faltas.columns),
        None,
    )
    custo_decisorio = valor_resumo("Custo decisório")
    if "Custo decisório" not in mapa_resumo:
        custo_decisorio = sum([
            valor_resumo("Receita não gerada", absoluto=True),
            valor_resumo("Contratação IS", absoluto=True),
            valor_resumo("Transporte por rota", absoluto=True),
            valor_resumo("Mobilização", absoluto=True),
            valor_resumo("Adicional noturno", absoluto=True),
        ])

    return {
        "status": "",
        "fo": custo_decisorio,
        "lucro": float("nan"),
        "receita_nao_gerada": valor_resumo("Receita não gerada", absoluto=True),
        "custo_contratacao": valor_resumo("Contratação IS", absoluto=True),
        "custo_transporte": valor_resumo("Transporte por rota", absoluto=True),
        "custo_mobilizacao": valor_resumo("Mobilização", absoluto=True),
        "mobilizacoes": valor_resumo("Mobilizações cobradas"),
        "custo_noturno": valor_resumo("Adicional noturno", absoluto=True),
        "contratacoes": (
            float(pd.to_numeric(contratacoes[coluna_contratacoes], errors="coerce").fillna(0).sum())
            if coluna_contratacoes else 0.0
        ),
        "faltas": (
            float(pd.to_numeric(faltas[coluna_faltas], errors="coerce").fillna(0).sum())
            if coluna_faltas else 0.0
        ),
    }


def tabela_comparacao(metricas_baseline, metricas_simulacao):
    definicoes = [
        ("Custo decisório", "fo", "menor", "moeda"),
        ("Receita não gerada", "receita_nao_gerada", "menor", "moeda"),
        ("Postos-dia não atendidos", "faltas", "menor", "numero"),
        ("Contratação de suplentes", "custo_contratacao", "menor", "moeda"),
        ("Suplentes contratados", "contratacoes", "menor", "numero"),
        ("Transporte", "custo_transporte", "menor", "moeda"),
        ("Mobilização", "custo_mobilizacao", "menor", "moeda"),
        ("Mobilizações cobradas", "mobilizacoes", "menor", "numero"),
        ("Adicional noturno", "custo_noturno", "menor", "moeda"),
    ]
    linhas = []
    for indicador, chave, melhor, formato in definicoes:
        base = float(metricas_baseline[chave])
        sim = float(metricas_simulacao[chave])
        delta = sim - base
        if abs(delta) < 1e-9:
            leitura = "Sem alteração"
        elif (melhor == "menor" and delta < 0) or (melhor == "maior" and delta > 0):
            leitura = "Melhorou"
        else:
            leitura = "Piorou"
        if formato == "moeda":
            base_fmt, sim_fmt, delta_fmt = _reais(base), _reais(sim), _reais(delta)
        else:
            base_fmt = f"{base:,.0f}".replace(",", ".")
            sim_fmt = f"{sim:,.0f}".replace(",", ".")
            delta_fmt = f"{delta:+,.0f}".replace(",", ".")
        linhas.append({
            "Indicador": indicador,
            "Baseline": base_fmt,
            "Simulação": sim_fmt,
            "Diferença (simulação - baseline)": delta_fmt,
            "Leitura": leitura,
        })
    return pd.DataFrame(linhas)


# =========================================================================
#                              INTERFACE
# =========================================================================
def render():
    aplicar_estilo_personalizado()

    if not usando_google():
        st.error(
            "A conexão com o Google ainda não foi configurada. No Streamlit Cloud, "
            "abra **Settings > Secrets** e cadastre o JSON da conta de serviço e "
            "os IDs das quatro planilhas."
        )
        st.info(
            "Use o arquivo `secrets_TEMPLATE.toml` do repositório como modelo. "
            "Não coloque o JSON real no GitHub."
        )
        return

    st.session_state.setdefault("cto_tela", None)

    def selecionar_tela(tela):
        if st.session_state.get("cto_tela") == tela:
            return
        st.session_state["cto_tela"] = tela
        st.session_state.pop("cto_plano", None)

    def voltar_ao_menu():
        st.session_state["cto_tela"] = None
        st.session_state.pop("cto_plano", None)
        st.session_state.pop("cto_solicitacoes_teste", None)

    if st.session_state["cto_tela"] is None:
        st.title("Agendamento de Férias CTO")
        st.markdown(
            "Escolha se deseja gerar o plano oficial de referência ou testar um cenário "
            "com solicitações de férias pendentes."
        )

        col_baseline, col_simulacao = st.columns(2)
        with col_baseline:
            st.button(
                "📋 Gerar baseline",
                type="primary",
                width="stretch",
                on_click=selecionar_tela,
                args=("baseline",),
            )
            st.info(
                "Gera o plano de férias, alocações e suplentes sem acrescentar "
                "solicitações pendentes. O resultado será salvo como referência."
            )
        with col_simulacao:
            st.button(
                "🧪 Simulação de férias",
                type="primary",
                width="stretch",
                on_click=selecionar_tela,
                args=("simulacao",),
            )
            st.info(
                "Permite selecionar solicitações pendentes, executar um cenário e "
                "comparar o resultado com o baseline."
            )
        return

    modo_simulacao = st.session_state["cto_tela"] == "simulacao"
    if modo_simulacao:
        st.title("Simulação de férias")
        st.caption(
            "Selecione solicitações pendentes para testar como aprovadas e compare o "
            "resultado com o baseline salvo."
        )
    else:
        st.title("Baseline — plano de férias e alocações")
        st.caption(
            "Execute o modelo sem solicitações pendentes adicionais. O resultado será "
            "salvo como referência para as simulações."
        )

    col_a, col_b = st.columns([3, 1])
    with col_a:
        st.button("← Voltar ao menu", on_click=voltar_ao_menu)
    with col_b:
        if st.button("🔄 Recarregar planilhas", width="stretch"):
            limpar_cache()
            carregar_solicitacoes_pendentes_base.clear()
            carregar_dados_base.clear()
            st.session_state.pop("cto_plano", None)
            st.session_state.pop("cto_solicitacoes_teste", None)
            st.rerun()

    solicitacoes_pendentes = []
    consideracoes_cadastro = []
    solicitacoes_por_id = {}
    solicitacoes_selecionadas_ids = []
    solicitacoes_para_modelo = ()

    def rotulo_solicitacao(solicitacao_id):
        item = solicitacoes_por_id[solicitacao_id]
        inicio = pd.Timestamp(item["inicio"]).strftime("%d/%m/%Y")
        fim = pd.Timestamp(item["fim"]).strftime("%d/%m/%Y")
        return f"{item['nome']} | Matrícula {item['matricula']} | {inicio} a {fim}"

    if modo_simulacao:
        try:
            carga_solicitacoes = carregar_solicitacoes_pendentes_base()
        except dados_ferias_cto.ErroValidacaoDados as e:
            st.error(str(e))
            return
        except Exception as e:
            st.error(f"Erro ao carregar as solicitações de férias pendentes: {e}")
            return

        solicitacoes_pendentes = carga_solicitacoes["solicitacoes"]
        consideracoes_cadastro = carga_solicitacoes.get("consideracoes", [])
        if consideracoes_cadastro:
            with st.expander(
                f"⚠️ Solicitações ignoradas por problemas de cadastro ({len(consideracoes_cadastro)})"
            ):
                for consideracao in consideracoes_cadastro:
                    st.warning(consideracao)

        solicitacoes_por_id = {item["id"]: item for item in solicitacoes_pendentes}

        st.caption(
            "Exibindo somente solicitações futuras de férias dos cargos considerados "
            "pelo modelo: auxiliar, laboratorista e técnico."
        )

        def limpar_plano_ao_mudar_cenario():
            st.session_state.pop("cto_plano", None)

        solicitacoes_selecionadas_ids = st.multiselect(
            "Solicitações de férias pendentes para testar como aprovadas",
            options=list(solicitacoes_por_id),
            format_func=rotulo_solicitacao,
            key="cto_solicitacoes_teste",
            on_change=limpar_plano_ao_mudar_cenario,
            help=(
                "Selecione as solicitações que devem ser consideradas aprovadas neste cenário. "
                "As solicitações não selecionadas serão ignoradas pelo modelo."
            ),
        )
        if not solicitacoes_pendentes:
            st.info("Não há solicitações de férias pendentes para testar.")

        solicitacoes_para_modelo = tuple(
            (
                solicitacoes_por_id[item_id]["matricula"],
                solicitacoes_por_id[item_id]["inicio"],
                solicitacoes_por_id[item_id]["fim"],
            )
            for item_id in solicitacoes_selecionadas_ids
        )

    texto_gerar = (
        "▶️ Executar simulação e comparar com o baseline"
        if modo_simulacao
        else "▶️ Gerar e salvar baseline (365 dias)"
    )
    gerar = st.button(texto_gerar, type="primary", width="stretch")

    if gerar:
        metricas_baseline = None
        if modo_simulacao:
            metricas_baseline = st.session_state.get("cto_baseline_metricas")
            if metricas_baseline is None:
                try:
                    metricas_baseline = carregar_metricas_baseline_local()
                except Exception as e:
                    st.error(f"Não foi possível ler o baseline salvo: {e}")
                    return
            if metricas_baseline is None:
                st.error(
                    "Nenhum baseline foi encontrado. Acesse **Gerar baseline**, execute o "
                    "modelo e salve a referência antes de rodar a simulação."
                )
                return

        st.session_state.pop("cto_plano", None)
        st.subheader("Andamento da execucao")
        progresso = st.status("Carregando dados e preparando o modelo...", expanded=True)

        # 1) Carrega os dados.
        try:
            dados, log_dados = carregar_dados_base(solicitacoes_para_modelo)
            progresso.update(label="Dados carregados. Resolvendo o modelo...", state="running")
        except dados_ferias_cto.ErroValidacaoDados as e:
            progresso.update(label="Dados de cadastro incompletos.", state="error")
            st.error(str(e))
            return
        except SystemExit as e:
            progresso.update(label="Carregamento interrompido.", state="error")
            st.error(f"O carregamento de dados foi interrompido: {e}")
            return
        except Exception as e:
            progresso.update(label="Erro ao carregar os dados.", state="error")
            st.error(f"Erro ao carregar os dados do modelo: {e}")
            return

        aviso_dist = extrair_aviso_distancia(log_dados)

        # 2) Resolve o modelo para o ano inteiro.
        try:
            with st.spinner("Resolvendo o modelo tático (pode levar alguns minutos)..."):
                resultados, status_str, log_modelo = rodar_modelo(
                    dados,
                    TIME_LIMIT,
                    GAP,
                    log_inicial=log_dados,
                )
            progresso.update(label=f"Modelo finalizado. Status: {status_str}", state="complete")
        except SystemExit as e:
            progresso.update(label="Modelo interrompido.", state="error")
            st.error(f"O modelo não montou um plano viável: {e}")
            return
        except Exception as e:
            progresso.update(label="Erro ao resolver o modelo.", state="error")
            st.error(f"Erro ao resolver o modelo: {e}")
            return

        metricas_plano = metricas(resultados, status_str)
        xlsx_plano = montar_excel_resultados(resultados)
        erro_salvar_resultado = None
        destino_resultado = None
        comparacao_baseline = None
        solucao_valida = status_str in ("Optimal", "Feasible")
        if modo_simulacao and solucao_valida:
            comparacao_baseline = tabela_comparacao(metricas_baseline, metricas_plano)
            try:
                destino_resultado = salvar_resultados_persistentes(
                    resultados,
                    modo="simulacao",
                    comparacao=comparacao_baseline,
                    conteudo_xlsx=xlsx_plano,
                )
            except Exception as e:
                erro_salvar_resultado = str(e)
        elif not modo_simulacao and solucao_valida:
            st.session_state["cto_baseline_metricas"] = metricas_plano
            try:
                destino_resultado = salvar_resultados_persistentes(
                    resultados,
                    modo="baseline",
                    conteudo_xlsx=xlsx_plano,
                )
            except Exception as e:
                erro_salvar_resultado = str(e)
        elif not modo_simulacao:
            erro_salvar_resultado = (
                f"o solver terminou com status {status_str}; um resultado não viável não pode ser usado como baseline"
            )

        st.session_state["cto_plano"] = {
            "modo": "simulacao" if modo_simulacao else "baseline",
            "status_str": status_str,
            "metricas": metricas_plano,
            "resultados": resultados,
            "xlsx": xlsx_plano,
            "aviso_dist": aviso_dist,
            "comparacao_baseline": comparacao_baseline,
            "erro_salvar_resultado": erro_salvar_resultado,
            "destino_resultado": destino_resultado,
            "solicitacoes_teste": [
                rotulo_solicitacao(item_id)
                for item_id in solicitacoes_selecionadas_ids
            ],
            "consideracoes_solicitacoes": (
                list(consideracoes_cadastro)
                + list(dados.get("consideracoes_solicitacoes", []))
            ),
            "solicitacoes_nao_selecionadas": max(
                len(solicitacoes_pendentes) - len(solicitacoes_selecionadas_ids), 0
            ),
        }

    # --- Renderização do último plano (sobrevive a reruns) ---
    plano = st.session_state.get("cto_plano")
    if not plano:
        st.info("Clique em **Gerar plano de alocação** para rodar o modelo.")
        return

    st.divider()

    plano_simulacao = plano.get("modo") == "simulacao"
    if plano_simulacao:
        if plano.get("solicitacoes_teste"):
            st.info(
                "**Solicitações selecionadas para serem testadas como aprovadas:**\n\n- "
                + "\n- ".join(plano["solicitacoes_teste"])
            )
        else:
            st.caption("Esta simulação não considerou nenhuma solicitação pendente como aprovada.")

        st.markdown("#### Considerações do cenário")
        consideracoes = plano.get("consideracoes_solicitacoes", [])
        if consideracoes:
            for consideracao in consideracoes:
                st.warning(consideracao)
        else:
            st.success("Nenhuma solicitação foi ignorada por inconsistência de cadastro ou período.")
        qtd_nao_selecionadas = plano.get("solicitacoes_nao_selecionadas", 0)
        if qtd_nao_selecionadas:
            st.caption(
                f"{qtd_nao_selecionadas} solicitação(ões) pendente(s) não selecionada(s) "
                "foram desconsideradas nesta simulação por escolha do usuário."
            )
    else:
        if plano.get("erro_salvar_resultado"):
            st.error(
                "O plano foi gerado, mas não foi possível salvar o baseline: "
                + plano["erro_salvar_resultado"]
            )
        else:
            st.success(f"Baseline salvo em {plano.get('destino_resultado') or NOME_BASELINE_LOCAL}.")

    if plano_simulacao:
        if plano.get("erro_salvar_resultado"):
            st.warning(
                "A simulação foi calculada, mas não foi possível salvá-la na planilha de resultados: "
                + plano["erro_salvar_resultado"]
            )
        elif plano.get("destino_resultado"):
            st.success("Simulação e comparação salvas na Planilha Google Resultados.")

    if plano.get("aviso_dist"):
        st.warning(
            "Pares de cidades ausentes na aba **Distancias** (calculados via API): "
            "adicione-os à aba.\n\n```\n" + plano["aviso_dist"] + "\n```"
        )

    if plano["status_str"] not in ("Optimal", "Feasible"):
        st.error(f"O solver terminou com status **{plano['status_str']}**. Os números podem não ser válidos.")
    else:
        st.success("Plano gerado com sucesso.")

    # --- Indicadores ---
    m = plano["metricas"]
    col1, col2, col3 = st.columns(3)
    col1.metric("Custo decisório", _reais(m["fo"]))
    col2.metric("Suplentes contratados", f"{m['contratacoes']:,.0f}".replace(",", "."))
    col3.metric("Postos-dia não atendidos", f"{m['faltas']:,.0f}".replace(",", "."))

    if plano_simulacao and plano.get("comparacao_baseline") is not None:
        st.subheader("Comparação com o baseline")
        st.dataframe(
            plano["comparacao_baseline"],
            width="stretch",
            hide_index=True,
        )

    # Aviso caso a regra de 14 dias seja violada (não deve acontecer: é restrição do modelo).
    violacoes = plano["resultados"].get("violacoes_bloco14", [])
    if violacoes:
        st.error(
            "⚠️ Regra de bloco de 14 dias possivelmente violada para: "
            + ", ".join(str(v) for v in violacoes)
            + ". Isso não deveria ocorrer — verifique o modelo."
        )

    # --- Tabelas ---
    r = plano["resultados"]
    t_resumo, t_ferias, t_contr, t_faltas, t_aloc, t_transp = st.tabs(
        ["Resumo", "Férias", "Contratações", "Faltas", "Alocações", "Transporte"]
    )
    with t_resumo:
        st.dataframe(_mostrar_streamlit(r["resumo"], "resumo"), width="stretch", hide_index=True)
        st.dataframe(_mostrar_streamlit(r["demanda"], "demanda"), width="stretch", hide_index=True)
    with t_ferias:
        df = r["ferias"]
        st.dataframe(df, width="stretch", hide_index=True) if not df.empty else st.info("Nenhum plano de férias.")
    with t_contr:
        df = r["contratacoes"]
        st.dataframe(_mostrar_streamlit(df, "contratacoes"), width="stretch", hide_index=True) if not df.empty else st.info("Nenhuma contratação de suplente.")
    with t_faltas:
        df = r["faltas"]
        st.dataframe(_mostrar_streamlit(df, "faltas"), width="stretch", hide_index=True) if not df.empty else st.info("Nenhum posto-dia sem atendimento.")
    with t_aloc:
        df = r["alocacoes"]
        if df.empty:
            st.info("Nenhuma alocação de suplente.")
        else:
            st.markdown("**Cobertura por perfil** — qual perfil cobre cada posto, agregado no período:")
            agg = (
                df.groupby(
                    ["tipo", "suplente", "nome_suplente", "rota_id", "projeto", "cargo_demanda", "turno"],
                    as_index=False,
                )
                .agg(
                    postos_dia=("quantidade_dias", "sum"),
                    primeira_data=("data_inicio", "min"),
                    ultima_data=("data_fim", "max"),
                )
                .sort_values(["projeto", "cargo_demanda", "postos_dia"], ascending=[True, True, False])
            )
            agg["postos_dia"] = agg["postos_dia"].round(0).astype(int)
            agg = agg.rename(columns={
                "tipo": "Tipo",
                "suplente": "Matrícula",
                "nome_suplente": "Nome",
                "rota_id": "Rota",
                "projeto": "Projeto",
                "cargo_demanda": "Cargo coberto",
                "turno": "Turno",
                "postos_dia": "Postos-dia",
                "primeira_data": "De",
                "ultima_data": "Até",
            })
            st.dataframe(agg, width="stretch", hide_index=True)

            with st.expander("Ver alocações detalhadas (dia a dia)"):
                st.dataframe(_mostrar_streamlit(df, "alocacoes"), width="stretch", hide_index=True)
    with t_transp:
        df = r["transporte"]
        st.dataframe(_mostrar_streamlit(df, "transporte"), width="stretch", hide_index=True) if not df.empty else st.info("Sem custo de transporte por mudança.")

    # --- Download ---
    st.download_button(
        "⬇️ Baixar resultado da simulação" if plano_simulacao else "⬇️ Baixar baseline",
        data=plano["xlsx"],
        file_name=("Simulacao_Resultados_Tatico.xlsx" if plano_simulacao else NOME_BASELINE_LOCAL),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
